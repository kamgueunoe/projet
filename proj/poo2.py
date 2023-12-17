import tkinter as tk
import sqlite3
from openpyxl import Workbook


# Connexion à la base de données
conn = sqlite3.connect('projets.db')
c = conn.cursor()

# Création de la table projets si elle n'existe pas
c.execute('''CREATE TABLE IF NOT EXISTS projets
             (id INTEGER PRIMARY KEY AUTOINCREMENT,
              nom TEXT,
              description TEXT,
              avancement INTEGER)''')

# Fonction pour ajouter un projet
def ajouter_projet():
    nom = nom_entry.get()
    description = desc_entry.get()
    avancement = int(avance_entry.get())
    c.execute("INSERT INTO projets (nom, description, avancement) VALUES (?, ?, ?)", (nom, description, avancement))
    conn.commit()
    actualiser_liste()

# Fonction pour supprimer un projet
def supprimer_projet():
    id_projet = int(project_list.curselection()[0]) + 1
    c.execute("DELETE FROM projets WHERE id=?", (id_projet,))
    conn.commit()
    actualiser_liste()

# Fonction pour modifier un projet
def modifier_projet():
    id_projet = int(project_list.curselection()[0]) + 1
    nom = nom_entry.get()
    description = desc_entry.get()
    avancement = int(avance_entry.get())
    c.execute("UPDATE projets SET nom=?, description=?, avancement=? WHERE id=?", (nom, description, avancement, id_projet))
    conn.commit()
    actualiser_liste()

# Fonction pour exporter les projets dans un fichier Excel
def exporter_projet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Projets"
    ws['A1'] = "ID"
    ws['B1'] = "Nom"
    ws['C1'] = "Description"
    ws['D1'] = "Avancement"
    for row in c.execute("SELECT * FROM projets"):
        ws.append(row)
    wb.save("projets.xlsx")

# Fonction pour afficher la liste des projets
def actualiser_liste():
    project_list.delete(0, tk.END)
    for row in c.execute("SELECT * FROM projets"):
        project_list.insert(tk.END, row[1])

# Créer une fenêtre tkinter
window = tk.Tk()

# Ajouter un titre à la fenêtre
window.title("Gestion de projets")

# Ajouter un label
label = tk.Label(window, text="Ajouter un projet")
label.pack()

# Ajouter un champ de saisie pour le nom du projet
nom_label = tk.Label(window, text="Nom du projet : ")
nom_label.pack()
nom_entry = tk.Entry(window)
nom_entry.pack()

# Ajouter un champ de saisie pour la description du projet
desc_label = tk.Label(window, text="Description : ")
desc_label.pack()
desc_entry = tk.Entry(window)
desc_entry.pack()

# Ajouter un champ de saisie pour l'avancement du projet
avance_label = tk.Label(window, text="Avancement : ")
avance_label.pack()
avance_entry = tk.Entry(window)
avance_entry.pack()

# Ajouter un bouton pour ajouter le projet
add_button = tk.Button(window, text="Ajouter", command=ajouter_projet)
add_button.pack()

# Ajouter un label pour la liste des projets
list_label = tk.Label(window, text="Liste des projets")
list_label.pack()

# Ajouter une liste pour afficher les projets
project_list = tk.Listbox(window)
project_list.pack()

# Ajouter un bouton pour supprimer un projet
delete_button = tk.Button(window, text="Supprimer", command=supprimer_projet)
delete_button.pack()

# Ajouter un bouton pour modifier un projet
modify_button = tk.Button(window, text="Modifier", command=modifier_projet)
modify_button.pack()

# Ajouter un bouton pour exporter les projets dans un fichier Excel
export_button = tk.Button(window, text="Exporter", command=exporter_projet)
export_button.pack()

# Afficher la fenêtre
window.mainloop()
