import tkinter as tk
import sqlite3
from openpyxl import Workbook


def page():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
    user = c.fetchone()
    if user:
        message_label.config(text="Connexion réussie", fg="green")
        
        # Connexion à la base de données
        conn = sqlite3.connect('projets.db')
        d = conn.cursor()

        # Création de la table projets si elle n'existe pas
        d.execute('''CREATE TABLE IF NOT EXISTS projets
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nom TEXT,
                    description TEXT,
                    avancement INTEGER)''')

        # Fonction pour ajouter un projet
        def ajouter_projet():
            nom = nom_entry.get()
            description = desc_entry.get()
            avancement = int(avance_entry.get())
            c.execute("INSERT INTO projets (nom, description, avancement) VALUES (?, ?, ?)", (nom, description, avancement))
            conn.commit()
            actualiser_liste()

        # Fonction pour supprimer un projet
        def supprimer_projet():
            id_projet = int(project_list.curselection()[0]) + 1
            c.execute("DELETE FROM projets WHERE id=?", (id_projet,))
            conn.commit()
            actualiser_liste()

        # Fonction pour modifier un projet
        def modifier_projet():
            id_projet = int(project_list.curselection()[0]) + 1
            nom = nom_entry.get()
            description = desc_entry.get()
            avancement = int(avance_entry.get())
            c.execute("UPDATE projets SET nom=?, description=?, avancement=? WHERE id=?", (nom, description, avancement, id_projet))
            conn.commit()
            actualiser_liste()

        # Fonction pour exporter les projets dans un fichier Excel
        def exporter_projet():
            wb = Workbook()
            ws = wb.active
            ws.title = "Projets"
            ws['A1'] = "ID"
            ws['B1'] = "Nom"
            ws['C1'] = "Description"
            ws['D1'] = "Avancement"
            for row in c.execute("SELECT * FROM projets"):
                ws.append(row)
            wb.save("projets.xlsx")

        # Fonction pour afficher la liste des projets
        def actualiser_liste():
            project_list.delete(0, tk.END)
            for row in c.execute("SELECT * FROM projets"):
                project_list.insert(tk.END, row[1])

        # Créer une fenêtre tkinter
        window = tk.Tk()

        # Ajouter un titre à la fenêtre
        window.title("Gestion de projets")

        # Ajouter un label
        label = tk.Label(window, text="Ajouter un projet")
        label.pack()

        # Ajouter un champ de saisie pour le nom du projet
        nom_label = tk.Label(window, text="Nom du projet : ")
        nom_label.pack()
        nom_entry = tk.Entry(window)
        nom_entry.pack()

        # Ajouter un champ de saisie pour la description du projet
        desc_label = tk.Label(window, text="Description : ")
        desc_label.pack()
        desc_entry = tk.Entry(window)
        desc_entry.pack()

        # Ajouter un champ de saisie pour l'avancement du projet
        avance_label = tk.Label(window, text="Avancement : ")
        avance_label.pack()
        avance_entry = tk.Entry(window)
        avance_entry.pack()

        # Ajouter un bouton pour ajouter le projet
        add_button = tk.Button(window, text="Ajouter", command=ajouter_projet)
        add_button.pack()

        # Ajouter un label pour la liste des projets
        list_label = tk.Label(window, text="Liste des projets")
        list_label.pack()

        # Ajouter une liste pour afficher les projets
        project_list = tk.Listbox(window)
        project_list.pack()

        # Ajouter un bouton pour supprimer un projet
        delete_button = tk.Button(window, text="Supprimer", command=supprimer_projet)
        delete_button.pack()

        # Ajouter un bouton pour modifier un projet
        modify_button = tk.Button(window, text="Modifier", command=modifier_projet)
        modify_button.pack()

        # Ajouter un bouton pour exporter les projets dans un fichier Excel
        export_button = tk.Button(window, text="Exporter", command=exporter_projet)
        export_button.pack()

        # Afficher la fenêtre
        window.mainloop()
    else:
        message_label.config(text="Nom d'utilisateur ou mot de passe incorrect", fg="red")
    

# Création de la fenêtre principale
root = tk.Tk()
root.title("Page de connexion")

# Connexion à la base de données
conn = sqlite3.connect("users.db")
c = conn.cursor()

# Création de la table des utilisateurs si elle n'existe pas
c.execute("""CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                username TEXT,
                password TEXT
            )""")
conn.commit()

# Fonction pour vérifier les informations de connexion
def verifier_connexion():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
    user = c.fetchone()
    if user:
        message_label.config(text="Connexion réussie", fg="green")
    else:
        message_label.config(text="Nom d'utilisateur ou mot de passe incorrect", fg="red")

# Fonction pour créer un nouveau compte
def creer_compte():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))
    conn.commit()
    message_label.config(text="Compte créé avec succès", fg="green")

# Création des widgets pour la page de connexion
username_label = tk.Label(root, text="Nom d'utilisateur")
username_label.grid(row=0, column=0)
username_entry = tk.Entry(root)
username_entry.grid(row=0, column=1)
password_label = tk.Label(root, text="Mot de passe")
password_label.grid(row=1, column=0)
password_entry = tk.Entry(root, show="*")
password_entry.grid(row=1, column=1)
connexion_button = tk.Button(root, text="Connexion", command=page)
connexion_button.grid(row=2, column=0)
creer_compte_button = tk.Button(root, text="Créer un compte", command=creer_compte)
creer_compte_button.grid(row=2, column=1)
message_label = tk.Label(root, text="")
message_label.grid(row=3, column=0, columnspan=2)

# Lancement de la boucle principale
root.mainloop()

# Fermeture de la connexion à la base de données
conn.close()
