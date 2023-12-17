import tkinter as tk
import sqlite3
from openpyxl import Workbook
from tkcalendar import DateEntry
from tkcalendar import Calendar

# Création de la base de données
conn = sqlite3.connect('aaa.db')
d = conn.cursor()

# Création de la table projets si elle n'existe pas
d.execute('''CREATE TABLE IF NOT EXISTS projets
    (id INTEGER PRIMARY KEY AUTOINCREMENT,
        tache TEXT, description TEXT, date TEXT)''')


def page1():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
    user = c.fetchone()
    if user:
        message_label.config(text="Connexion réussie", fg="green")

        # Fonction pour ajouter une tache
        def ajouter_tache():
            tache = tache_entry.get()
            description = desc_entry.get()
            date = cal.get_date()

            if tache == '' or description == '' or date == '':
                return

            else:
                # Ajout du projet à la base de données
                d.execute('''INSERT INTO projets (tache, description, date) VALUES (?, ?, ?)''', (tache, description, date))
                conn.commit()

                actualiser_liste()

        # Fonction pour supprimer un projet
        def supprimer_tache():
            id_projet = int(project_list.curselection()[0]) + 1
            d.execute("DELETE FROM projets WHERE id=?", (id_projet,))
            conn.commit()
            actualiser_liste()

        # Fonction pour modifier un projet
        def modifier_tache():
            id_projet = int(project_list.curselection()[0]) + 1
            tache = tache_entry.get()
            description = desc_entry.get()
            date = cal.get_date()  
            d.execute("UPDATE projets SET tache=?, description=?, date=? WHERE id=?", (tache, description, date, id_projet))
            conn.commit()
            actualiser_liste()

        # Fonction pour exporter les projets dans un fichier Excel
        def exporter_tache():
            wb = Workbook()
            ws = wb.active
            ws.title = "Projets"
            ws['A1'] = "ID"
            ws['B1'] = "tache"
            ws['C1'] = "Description"
            ws['D1'] = "Date"
            for row in d.execute("SELECT * FROM projets"):
                ws.append(row)
            wb.save("projets.xlsx")

        # Fonction pour afficher la liste des projets
        def actualiser_liste():
            tache = tache_entry.get()
            description = desc_entry.get()
            date = cal.get_date() 
            project_list.insert(tk.END, f"{tache} - {description}- {date}" )
            tache_entry.delete(0, tk.END)
            desc_entry.delete(0, tk.END)

        # Créer une fenêtre tkinter
        window = tk.Tk()

        # Ajouter un titre à la fenêtre
        window.title("Gestion de projets")

        # Ajouter un label
        label = tk.Label(window, text="Ajouter une tache")
        label.pack()

        # Ajouter un champ de saisie pour la tache du projet
        tache_label = tk.Label(window, text="tache du projet : ")
        tache_label.pack()
        tache_entry = tk.Entry(window)
        tache_entry.pack()

        # Ajouter un champ de saisie pour la description de la tache
        desc_label = tk.Label(window, text="Description : ")
        desc_label.pack()
        desc_entry = tk.Entry(window)
        desc_entry.pack()

        # Création du champ de saisie de date
        cal = DateEntry(window, width=12, background='darkblue',
                        foreground='white', borderwidth=2)
        cal.pack()



        # Ajouter un bouton pour ajouter la tache
        add_button = tk.Button(window, text="Ajouter", command=ajouter_tache)
        add_button.pack()

        # Ajouter un label pour la liste des taches
        list_label = tk.Label(window, text="Liste des taches")
        list_label.pack()

        # Ajouter une liste pour afficher les taches
        project_list = tk.Listbox(window, width=100)
        project_list.pack()

        # Ajouter un bouton pour supprimer une tache
        delete_button = tk.Button(window, text="Supprimer", command=supprimer_tache)
        delete_button.pack()

        # Ajouter un bouton pour modifier une tache
        modify_button = tk.Button(window, text="Modifier", command=modifier_tache)
        modify_button.pack()

        # Ajouter un bouton pour ouvrir le calendrier des objectifs
        modify_button = tk.Button(window, text="Ouvrir le calendrier des objectifs", command=page2)
        modify_button.pack()

        # Ajouter un bouton pour exporter les projets dans un fichier Excel
        export_button = tk.Button(window, text="Exporter", command=exporter_tache)
        export_button.pack()

                


        # Afficher la fenêtre
        window.mainloop()
    else:
        message_label.config(text="nom d'utilisateur ou mot de passe incorrect", fg="red")
    
def page2():
    root = tk.Tk()
    root.title("Calendrier et objectifs du projet")

    # Fonction pour afficher le calendrier
    def afficher_calendrier():
        top = tk.Toplevel(root)
        cal = Calendar(top, font="Arial 14", selectmode='day', locale='fr_FR')
        cal.pack(fill="both", expand=True)

    # Création des widgets pour afficher le calendrier
    calendrier_label = tk.Label(root, text="Calendrier")
    calendrier_label.grid(row=0, column=0)
    calendrier_button = tk.Button(root, text="Afficher calendrier", command=afficher_calendrier)
    calendrier_button.grid(row=0, column=1)

    # Fonction pour ajouter un objectif dans la listebox
    def ajouter_objectif():
        objectif = objectif_entry.get()
        date = date_entry.get()
        objectif_listbox.insert(tk.END, f"{date} - {objectif}")
        objectif_entry.delete(0, tk.END)
        date_entry.delete(0, tk.END)

    # Fonction pour supprimer un objectif de la listebox
    def supprimer_objectif():
        selection = objectif_listbox.curselection()
        if selection:
            objectif_listbox.delete(selection)

    # Création des widgets pour ajouter et supprimer des objectifs
    objectifs_label = tk.Label(root, text="Objectifs")
    objectifs_label.grid(row=1, column=0)
    objectif_entry = tk.Entry(root)
    objectif_entry.grid(row=1, column=1)
    date_label = tk.Label(root, text="Date")
    date_label.grid(row=2, column=0)
    date_entry = tk.Entry(root)
    date_entry.grid(row=2, column=1)
    ajouter_objectif_button = tk.Button(root, text="Ajouter objectif", command=ajouter_objectif)
    ajouter_objectif_button.grid(row=3, column=0)
    supprimer_objectif_button = tk.Button(root, text="Supprimer objectif", command=supprimer_objectif)
    supprimer_objectif_button.grid(row=3, column=1)

    # Création de la listebox pour afficher les objectifs
    objectif_listbox = tk.Listbox(root, width=100)
    objectif_listbox.grid(row=4, column=0, columnspan=2)

    # Ajout d'exemples d'objectifs
    objectif_listbox.insert(tk.END, "2022-01-15 - Finaliser le design de l'interface utilisateur")
    objectif_listbox.insert(tk.END, "2022-02-01 - Terminer la phase de développement")

    # Lancement de la boucle principale
    root.mainloop()


# Création de la fenêtre principale
root = tk.Tk()
root.title("Page de connexion")

# Connexion à la base de données
conn = sqlite3.connect("users.db")
c = conn.cursor()

# Création de la table des utilisateurs si elle n'existe pas
c.execute("""CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                username TEXT,
                password TEXT
            )""")
conn.commit()

# Fonction pour vérifier les informations de connexion
def verifier_connexion():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
    user = c.fetchone()
    if user:
        message_label.config(text="Connexion réussie", fg="green")
    else:
        message_label.config(text="Nom d'utilisateur ou mot de passe incorrect", fg="red")

# Fonction pour créer un nouveau compte
def creer_compte():
    username = username_entry.get()
    password = password_entry.get()
    c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))
    conn.commit()
    message_label.config(text="Compte créé avec succès", fg="green")

# Création des widgets pour la page de connexion
username_label = tk.Label(root, text="Nom d'utilisateur")
username_label.pack()
username_entry = tk.Entry(root)
username_entry.pack()
password_label = tk.Label(root, text="Mot de passe")
password_label.pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()
connexion_button = tk.Button(root, text="Connexion", command=page1)
connexion_button.pack()
creer_compte_button = tk.Button(root, text="Créer un compte", command=creer_compte)
creer_compte_button.pack()
message_label = tk.Label(root, text="")
message_label.pack()

# Lancement de la boucle principale
root.mainloop()

# Fermeture de la connexion à la base de données
conn.close()
